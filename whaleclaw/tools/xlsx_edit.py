"""XLSX edit tool — targeted cell/text replacement in an existing .xlsx file."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from whaleclaw.tools.base import Tool, ToolDefinition, ToolParameter, ToolResult
from whaleclaw.tools.deps import ensure_tool_dep


class XlsxEditTool(Tool):
    """Edit an existing Excel workbook by cell or global text replacement."""

    @property
    def definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="xlsx_edit",
            description="修改现有 Excel（.xlsx）内容，可指定工作表与单元格，避免整表重做。",
            parameters=[
                ToolParameter(name="path", type="string", description="XLSX 文件绝对路径"),
                ToolParameter(
                    name="sheet",
                    type="string",
                    description="工作表名称（可选，不填默认首个）",
                    required=False,
                ),
                ToolParameter(
                    name="cell",
                    type="string",
                    description="单元格地址（可选，如 A1）",
                    required=False,
                ),
                ToolParameter(name="old_text", type="string", description="要替换的原文"),
                ToolParameter(name="new_text", type="string", description="替换后的新文案"),
            ],
        )

    async def execute(self, **kwargs: Any) -> ToolResult:
        if not ensure_tool_dep("openpyxl"):
            return ToolResult(success=False, output="", error="缺少依赖 openpyxl")

        from openpyxl import load_workbook
        from openpyxl.cell.cell import Cell

        raw_path = str(kwargs.get("path", "")).strip()
        sheet_name = str(kwargs.get("sheet", "")).strip()
        cell_ref = str(kwargs.get("cell", "")).strip().upper()
        old_text = str(kwargs.get("old_text", "")).strip()
        new_text = str(kwargs.get("new_text", ""))

        if not raw_path:
            return ToolResult(success=False, output="", error="path 不能为空")
        if not old_text:
            return ToolResult(success=False, output="", error="old_text 不能为空")

        path = Path(raw_path).expanduser().resolve()
        if not path.is_file():
            return ToolResult(success=False, output="", error=f"文件不存在: {path}")
        if path.suffix.lower() != ".xlsx":
            return ToolResult(success=False, output="", error="仅支持 .xlsx 文件")

        try:
            wb = load_workbook(filename=str(path))
        except Exception as exc:
            return ToolResult(success=False, output="", error=f"XLSX 打开失败: {exc}")

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return ToolResult(success=False, output="", error=f"工作表不存在: {sheet_name}")
            sheets = [wb[sheet_name]]
        else:
            sheets = [wb[wb.sheetnames[0]]]

        replaced = 0
        touched_cells = 0

        if cell_ref:
            ws = sheets[0]
            value = ws[cell_ref].value
            if not isinstance(value, str):
                return ToolResult(
                    success=False,
                    output="",
                    error=f"{ws.title}!{cell_ref} 不是文本单元格",
                )
            count = value.count(old_text)
            if count <= 0:
                return ToolResult(
                    success=False,
                    output="",
                    error=f"{ws.title}!{cell_ref} 未找到匹配文本",
                )
            ws[cell_ref].value = value.replace(old_text, new_text)
            replaced = count
            touched_cells = 1
        else:
            for ws in sheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if not isinstance(cell, Cell):
                            continue
                        if not isinstance(cell.value, str):
                            continue
                        count = cell.value.count(old_text)
                        if count <= 0:
                            continue
                        cell.value = cell.value.replace(old_text, new_text)
                        replaced += count
                        touched_cells += 1

        if replaced == 0:
            return ToolResult(success=False, output="", error="工作簿中未找到匹配文本")

        try:
            wb.save(str(path))
        except Exception as exc:
            return ToolResult(success=False, output="", error=f"XLSX 保存失败: {exc}")

        return ToolResult(
            success=True,
            output=f"已修改 {path}，替换 {replaced} 处文本，影响 {touched_cells} 个单元格",
        )
